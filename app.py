"""
FeedGen - AI Feed Text Generator
=================================
Web service for generating optimized product titles and descriptions
for Google Merchant Center feeds using Claude AI.

Usage:
    pip install fastapi uvicorn anthropic openpyxl python-multipart
    uvicorn app:app --host 0.0.0.0 --port 8000
"""

import os
import json
import re
import uuid
import asyncio
import tempfile
from pathlib import Path
from datetime import datetime

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.requests import Request
from fastapi.responses import FileResponse, JSONResponse

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = FastAPI(title="FeedGen", version="1.0")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Storage for job status
jobs = {}
UPLOAD_DIR = Path(tempfile.gettempdir()) / "feedgen"
UPLOAD_DIR.mkdir(exist_ok=True)


# =============================================================================
# ROUTES
# =============================================================================

@app.get("/")
async def index(request: Request):
    return templates.TemplateResponse(request=request, name="index.html")


@app.post("/api/analyze")
async def analyze_feed(file: UploadFile = File(...)):
    """Аналізує завантажений xlsx і повертає структуру фіду."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Підтримуються лише .xlsx файли")

    tmp_path = UPLOAD_DIR / f"{uuid.uuid4().hex}_{file.filename}"
    content = await file.read()
    tmp_path.write_bytes(content)

    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Detect header
        has_header = isinstance(rows[0][0], str) and rows[0][0].lower() in ['id', 'item_id']
        if has_header:
            headers = [str(h) for h in rows[0]]
            data_rows = rows[1:]
        else:
            headers = [f"col_{i}" for i in range(len(rows[0]))]
            data_rows = rows

        from collections import Counter
        titles = Counter(row[1] for row in data_rows if len(row) > 1 and row[1])

        # Sample values for key columns
        sample = {}
        col_indices = {'title': 1, 'product_type': 7, 'gender': 18, 'color': 21, 'material': 22}
        for name, idx in col_indices.items():
            vals = set()
            for row in data_rows[:100]:
                if len(row) > idx and row[idx]:
                    vals.add(str(row[idx]))
            sample[name] = sorted(vals)[:10]

        return {
            "file_id": tmp_path.stem,
            "filename": file.filename,
            "total_rows": len(data_rows),
            "unique_titles": len(titles),
            "has_header": has_header,
            "columns": headers[:28],
            "sample_values": sample,
        }
    except Exception as e:
        raise HTTPException(400, f"Помилка читання файлу: {str(e)}")


@app.post("/api/generate")
async def start_generation(
    background_tasks: BackgroundTasks,
    file_id: str = Form(...),
    config: UploadFile = File(None),
    config_text: str = Form(None),
    model: str = Form("claude-sonnet-4-20250514"),
    language: str = Form("uk"),
):
    """Запускає генерацію в фоновому режимі."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(500, "ANTHROPIC_API_KEY не налаштований на сервері")

    # Find uploaded file
    matching = list(UPLOAD_DIR.glob(f"{file_id}*"))
    if not matching:
        raise HTTPException(404, "Файл не знайдено. Завантажте ще раз.")
    input_path = matching[0]

    # Parse config
    niche_config = None
    if config:
        niche_config = json.loads(await config.read())
    elif config_text:
        niche_config = json.loads(config_text)

    job_id = uuid.uuid4().hex[:12]
    jobs[job_id] = {
        "status": "processing",
        "progress": 0,
        "total": 0,
        "message": "Підготовка...",
        "created": datetime.now().isoformat(),
        "output_file": None,
        "errors": [],
    }

    background_tasks.add_task(
        run_generation, job_id, str(input_path), api_key, model, language, niche_config
    )

    return {"job_id": job_id}


@app.get("/api/status/{job_id}")
async def get_status(job_id: str):
    if job_id not in jobs:
        raise HTTPException(404, "Завдання не знайдено")
    return jobs[job_id]


@app.get("/api/download/{job_id}")
async def download_result(job_id: str):
    if job_id not in jobs:
        raise HTTPException(404, "Завдання не знайдено")
    job = jobs[job_id]
    if job["status"] != "done" or not job["output_file"]:
        raise HTTPException(400, "Файл ще не готовий")
    return FileResponse(
        job["output_file"],
        filename=f"generated_feed_{job_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =============================================================================
# GENERATION ENGINE
# =============================================================================

def build_system_prompt(niche_config, language):
    """Будує системний промпт з конфігу ніші або дефолтний."""
    if niche_config:
        return json.dumps(niche_config, ensure_ascii=False, indent=2)

    # Default system prompt for generic feed generation
    return f"""Ти — спеціалізований ШІ-копірайтер для товарних фідів Google Merchant Center.
Мова генерації: {"українська" if language == "uk" else "English"}.

Твоя задача — на основі наданих атрибутів товару згенерувати:
1. Оптимізований title (до 150 символів) — з ключовими характеристиками товару для максимальної пошукової релевантності
2. Оптимізований description (до 500 слів) — структурований опис з характеристиками, перевагами, матеріалами

Правила:
- Не вигадуй характеристики, яких немає в даних
- Title: найважливіше в перших 70 символах
- Природна мова, без спаму ключовиками
- Кожен title має бути унікальним

Відповідай ТІЛЬКИ у форматі JSON:
{{"title": "...", "description": "..."}}"""


def extract_product_data(row, has_header):
    """Витягує дані товару з рядка фіду."""
    return {
        "id": row[0],
        "title": row[1] or '',
        "description": str(row[2])[:500] if row[2] else '',
        "product_type": row[7] if len(row) > 7 else '',
        "brand": row[16] if len(row) > 16 else '',
        "gender": row[18] if len(row) > 18 else '',
        "color": row[21] if len(row) > 21 else '',
        "material": row[22] if len(row) > 22 else '',
    }


async def call_claude(client, model, system_prompt, product_data):
    """Один виклик Claude API для генерації title + description."""
    user_message = f"""Згенеруй оптимізований title і description для цього товару.

Поточний title: {product_data['title']}
Опис: {product_data['description'][:300]}
Product type: {product_data['product_type']}
Brand: {product_data['brand']}
Gender: {product_data['gender']}
Color: {product_data['color']}
Material: {product_data['material']}

ВАЖЛИВО: Відповідай ВИКЛЮЧНО валідним JSON без будь-якого іншого тексту.
Формат: {{"title": "новий заголовок", "description": "новий опис"}}"""

    try:
        response = client.messages.create(
            model=model,
            max_tokens=1024,
            temperature=0.3,
            system=system_prompt,
            messages=[{"role": "user", "content": user_message}],
        )

        text = response.content[0].text.strip()

        # Strategy 1: direct JSON parse
        try:
            result = json.loads(text)
            if "title" in result:
                return result["title"], result.get("description", "")
        except json.JSONDecodeError:
            pass

        # Strategy 2: extract from ```json ... ``` blocks
        code_block = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
        if code_block:
            try:
                result = json.loads(code_block.group(1))
                if "title" in result:
                    return result["title"], result.get("description", "")
            except json.JSONDecodeError:
                pass

        # Strategy 3: find first { ... } in text
        brace_match = re.search(r'\{[^{}]*"title"[^{}]*\}', text, re.DOTALL)
        if brace_match:
            try:
                result = json.loads(brace_match.group(0))
                if "title" in result:
                    return result["title"], result.get("description", "")
            except json.JSONDecodeError:
                pass

        # Strategy 4: find nested JSON (with inner braces in description)
        first_brace = text.find('{')
        last_brace = text.rfind('}')
        if first_brace != -1 and last_brace > first_brace:
            try:
                result = json.loads(text[first_brace:last_brace + 1])
                if "title" in result:
                    return result["title"], result.get("description", "")
            except json.JSONDecodeError:
                pass

        # All strategies failed
        return None, f"Could not parse response: {text[:100]}"

    except Exception as e:
        return None, str(e)


async def run_generation(job_id, input_path, api_key, model, language, niche_config):
    """Фонова задача генерації."""
    import anthropic

    try:
        client = anthropic.Anthropic(api_key=api_key)
        system_prompt = build_system_prompt(niche_config, language)

        # Read input
        wb = openpyxl.load_workbook(input_path, read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        has_header = isinstance(all_rows[0][0], str) and all_rows[0][0].lower() in ['id', 'item_id']
        data_rows = all_rows[1:] if has_header else all_rows

        # Deduplicate by (title, description)
        unique_products = {}
        row_keys = []
        for row in data_rows:
            key = (row[1], row[2])
            if key not in unique_products:
                unique_products[key] = extract_product_data(row, has_header)
            row_keys.append(key)

        total_unique = len(unique_products)
        jobs[job_id]["total"] = total_unique
        jobs[job_id]["message"] = f"Генерація {total_unique} унікальних товарів..."

        # Generate for unique products
        generated = {}
        processed = 0

        for key, product in unique_products.items():
            gen_title, gen_desc = await call_claude(client, model, system_prompt, product)

            if gen_title is None:
                gen_title = product['title']
                gen_desc = str(product.get('description', ''))
                jobs[job_id]["errors"].append(f"ID {product['id']}: fallback to original")

            generated[key] = (gen_title, gen_desc)
            processed += 1
            jobs[job_id]["progress"] = processed
            jobs[job_id]["message"] = f"Оброблено {processed}/{total_unique} товарів"

            # Rate limiting - 1 request per second to be safe
            await asyncio.sleep(0.5)

        # Build output
        output_rows = []
        for i, row in enumerate(data_rows):
            key = row_keys[i]
            g = generated[key]
            rid = int(row[0]) if row[0] else ''
            output_rows.append((rid, g[0], row[1], g[1], row[2]))

        # Write xlsx
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Generated Feed"

        out_headers = ['id', 'generated_title', 'title', 'generated_description', 'description']
        hfont = Font(bold=True, size=11, name='Arial')
        hfill = PatternFill('solid', start_color='D9E1F2')

        for ci, h in enumerate(out_headers, 1):
            c = out_ws.cell(row=1, column=ci, value=h)
            c.font, c.fill = hfont, hfill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for ri, rd in enumerate(output_rows, 2):
            for ci, val in enumerate(rd, 1):
                c = out_ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name='Arial', size=10)
                c.alignment = Alignment(vertical='top', wrap_text=True)

        widths = {'A': 10, 'B': 55, 'C': 45, 'D': 70, 'E': 50}
        for col, w in widths.items():
            out_ws.column_dimensions[col].width = w
        out_ws.auto_filter.ref = f"A1:E{len(output_rows)+1}"

        output_path = UPLOAD_DIR / f"generated_{job_id}.xlsx"
        out_wb.save(str(output_path))

        jobs[job_id]["status"] = "done"
        jobs[job_id]["output_file"] = str(output_path)
        jobs[job_id]["message"] = f"Готово! {len(output_rows)} товарів оброблено."

    except Exception as e:
        jobs[job_id]["status"] = "error"
        jobs[job_id]["message"] = f"Помилка: {str(e)}"


# =============================================================================
# HEALTH
# =============================================================================

@app.get("/api/health")
async def health():
    has_key = bool(os.environ.get("ANTHROPIC_API_KEY"))
    return {"status": "ok", "api_key_configured": has_key}
